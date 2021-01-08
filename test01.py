import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

excel_test = load_workbook(filename='./NewsResult.xlsx')
word_sheet = excel_test['sheet']

putdocID = word_sheet.cell(4, 1).value  # row, column
print(putdocID, '\n')

url = 'https://www.bigkinds.or.kr/news/detailView.do?docId=' + putdocID + '&returnCnt=1&sectionDiv=1000'
print(url)

# original url = 'https://www.bigkinds.or.kr/news/detailView.do?docId=02100201.20200707103441001&returnCnt=1&sectionDiv=1000'

response = requests.get(url)
html = response.text
print(html)

# 받아온 문자열에서 특정 문자열 위치 찾기
search = '"CONTENT":"'
indexNo = html.find(search)
print(indexNo, '\n')
indexNo += 11  # "CONTENT":" 문자수 = 11

# 태그 제거
cleanText = BeautifulSoup(html[indexNo:], 'lxml').text
print(cleanText)

